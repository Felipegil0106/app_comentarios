"""Pruebas rapidas de la logica interna (no abre el navegador ni la ventana).

Ejecutar con:   .venv\\Scripts\\python.exe pruebas.py

Sirve para comprobar que la limpieza de texto, la interpretacion de fechas,
la base de datos y la exportacion funcionan bien despues de cualquier cambio.
"""

from __future__ import annotations

import sys
import tempfile
from datetime import datetime, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from src.core.base_datos import BaseDatos            # noqa: E402
from src.core.exportar import exportar_csv, exportar_excel  # noqa: E402
from src.core.fechas import desde_epoch, interpretar_fecha  # noqa: E402
from src.core.limpieza import limpiar_autor, limpiar_texto  # noqa: E402
from src.core.modelos import Comentario, Publicacion  # noqa: E402

fallos: list[str] = []


def comprobar(condicion: bool, descripcion: str) -> None:
    estado = "OK  " if condicion else "FALLO"
    print(f"  [{estado}] {descripcion}")
    if not condicion:
        fallos.append(descripcion)


# --------------------------------------------------------------------- fechas
def probar_fechas() -> None:
    print("\n== Interpretacion de fechas ==")
    ahora = datetime(2025, 8, 20, 15, 0)

    f = interpretar_fecha("Hace 3 h", ahora)
    comprobar(f is not None and f.hour == 12, "'Hace 3 h' -> 12:00 del mismo dia")

    f = interpretar_fecha("2 h", ahora)
    comprobar(f is not None and f.hour == 13, "'2 h' -> 13:00")

    f = interpretar_fecha("Ayer a las 14:30", ahora)
    comprobar(
        f is not None and f.day == 19 and f.hour == 14 and f.minute == 30,
        "'Ayer a las 14:30' -> 19/08 14:30",
    )

    f = interpretar_fecha("3 de agosto de 2024 a las 10:15", ahora)
    comprobar(
        f == datetime(2024, 8, 3, 10, 15), "'3 de agosto de 2024 a las 10:15'"
    )

    f = interpretar_fecha("31 dic 2023", ahora)
    comprobar(f is not None and f.month == 12 and f.day == 31 and f.year == 2023,
              "'31 dic 2023'")

    f = interpretar_fecha("5 ago", ahora)
    comprobar(f is not None and f.year == 2025 and f.month == 8 and f.day == 5,
              "'5 ago' sin año -> año actual")

    # Sin año y en el futuro => debe asumir el año anterior
    f = interpretar_fecha("20 de diciembre", ahora)
    comprobar(f is not None and f.year == 2024, "'20 de diciembre' futuro -> 2024")

    f = interpretar_fecha("August 3, 2024 at 10:15 AM", ahora)
    comprobar(f == datetime(2024, 8, 3, 10, 15), "Formato en ingles")

    f = interpretar_fecha("2 semanas", ahora)
    comprobar(f is not None and (ahora - f).days == 14, "'2 semanas'")

    comprobar(interpretar_fecha("Me gusta", ahora) is None, "Texto sin fecha -> None")

    f = desde_epoch(1723456789)
    comprobar(f is not None and f.year == 2024, "creation_time de Facebook")
    comprobar(desde_epoch(123) is None, "Timestamp absurdo -> None")


# -------------------------------------------------------------------- limpieza
def probar_limpieza() -> None:
    print("\n== Limpieza de texto (solo texto + emojis) ==")

    comprobar(
        limpiar_texto("Me encanto este producto 😍🔥") == "Me encanto este producto 😍🔥",
        "Conserva texto con emojis",
    )
    comprobar(limpiar_texto("😍🔥👏") == "😍🔥👏", "Un comentario de solo emojis se conserva")
    comprobar(limpiar_texto("GIF") == "", "Un GIF suelto se descarta")
    comprobar(limpiar_texto("  Sticker  ") == "", "Un sticker suelto se descarta")
    comprobar(
        limpiar_texto("Ana envio un archivo adjunto") == "",
        "'envio un archivo adjunto' se descarta",
    )
    comprobar(limpiar_texto("...") == "", "Solo puntuacion -> vacio")
    comprobar(
        limpiar_texto("Buenisimo​​   trabajo") == "Buenisimo trabajo",
        "Quita caracteres invisibles y espacios de sobra",
    )
    comprobar(
        limpiar_texto("Excelente Me gusta") == "Excelente",
        "Quita el boton 'Me gusta' pegado al final",
    )
    comprobar(
        limpiar_texto("Foto\nQue lindo lugar") == "Que lindo lugar",
        "Quita la linea 'Foto' pero conserva el texto real",
    )
    comprobar(limpiar_autor("Juan Perez Top fan") == "Juan Perez", "Limpia 'Top fan'")
    comprobar(limpiar_autor("") == "(desconocido)", "Autor vacio -> (desconocido)")


# ---------------------------------------------------------------- base de datos
def probar_base_datos() -> Path:
    print("\n== Base de datos ==")
    tmp = Path(tempfile.mkdtemp()) / "prueba.sqlite3"
    bd = BaseDatos(tmp)

    hoy = datetime.now()
    pub1 = Publicacion(
        url="https://www.facebook.com/pagina/posts/1",
        red="facebook", fecha=hoy - timedelta(days=2), tipo="publicacion",
        texto="Lanzamiento del producto",
    )
    pub2 = Publicacion(
        url="https://www.facebook.com/reel/2",
        red="facebook", fecha=hoy - timedelta(days=5), tipo="reel",
        texto="Detras de camaras",
    )
    bd.guardar_publicacion(pub1)
    bd.guardar_publicacion(pub2)
    comprobar(len(bd.publicaciones()) == 2, "Se guardaron 2 publicaciones")

    # Guardar la misma publicacion otra vez no debe duplicar
    bd.guardar_publicacion(pub1)
    comprobar(len(bd.publicaciones()) == 2, "Re-guardar no duplica publicaciones")

    comentarios = [
        Comentario(pub1.url, "Ana", "Que buen producto 😍", red="facebook"),
        Comentario(pub1.url, "Luis", "Cuanto cuesta?", red="facebook"),
        Comentario(pub1.url, "Ana", "Ya lo compre", red="facebook", es_respuesta=True),
        Comentario(pub2.url, "Marta", "Me encanta este reel 🔥", red="facebook"),
    ]
    nuevos = bd.guardar_comentarios(comentarios)
    comprobar(nuevos == 4, "Se guardaron 4 comentarios nuevos")

    repetidos = bd.guardar_comentarios(comentarios)
    comprobar(repetidos == 0, "Volver a guardar los mismos no duplica")

    bd.recalcular_conteos()
    est = bd.estadisticas()
    comprobar(est["comentarios_totales"] == 4, "Total de comentarios = 4")
    comprobar(est["publicaciones_con_comentarios"] == 2, "Publicaciones con comentarios = 2")
    comprobar(est["autores_unicos"] == 3, "Autores unicos = 3")
    comprobar(est["respuestas"] == 1, "Respuestas = 1")

    # Filtro por URL: el contador debe bajar
    solo_pub1 = bd.comentarios(url=pub1.url)
    comprobar(len(solo_pub1) == 3, "Filtrar por URL de pub1 -> 3 comentarios")
    solo_pub2 = bd.comentarios(url=pub2.url)
    comprobar(len(solo_pub2) == 1, "Filtrar por URL de pub2 -> 1 comentario")

    # Busqueda por texto y por autor
    comprobar(len(bd.comentarios(busqueda="reel")) == 1, "Buscar 'reel' -> 1")
    comprobar(len(bd.comentarios(autor="Ana")) == 2, "Filtrar autor 'Ana' -> 2")
    comprobar(len(bd.comentarios(solo_respuestas=True)) == 1, "Solo respuestas -> 1")

    pares = bd.urls_con_comentarios()
    comprobar(len(pares) == 2 and pares[0][1] == 3, "urls_con_comentarios ordenado")

    # Los datos de la publicacion viajan junto al comentario
    fila = bd.comentarios(url=pub2.url)[0]
    comprobar(fila.get("tipo_publicacion") == "reel", "El JOIN trae el tipo de publicacion")

    print("\n== Exportacion ==")
    filas = bd.comentarios()
    destino_csv = tmp.parent / "salida.csv"
    exportar_csv(filas, destino_csv)
    contenido = destino_csv.read_text(encoding="utf-8-sig")
    comprobar(destino_csv.exists() and destino_csv.stat().st_size > 0, "CSV creado")
    comprobar("😍" in contenido, "El CSV conserva los emojis")
    comprobar("Comentario" in contenido.splitlines()[0], "El CSV tiene cabecera")

    destino_xlsx = tmp.parent / "salida.xlsx"
    exportar_excel(filas, destino_xlsx)
    comprobar(destino_xlsx.exists() and destino_xlsx.stat().st_size > 0, "Excel creado")

    import openpyxl
    libro = openpyxl.load_workbook(destino_xlsx)
    comprobar(
        libro.sheetnames == ["Comentarios", "Resumen por publicacion", "Top autores"],
        "El Excel tiene las 3 hojas",
    )
    comprobar(libro["Comentarios"].max_row == 5, "El Excel tiene 4 filas + cabecera")

    # Exportar una lista vacia no debe reventar
    exportar_csv([], tmp.parent / "vacio.csv")
    exportar_excel([], tmp.parent / "vacio.xlsx")
    comprobar(True, "Exportar vacio no falla")

    bd.cerrar()
    return tmp.parent


# -------------------------------------------------------------- URLs y modelos
def probar_facebook_urls() -> None:
    print("\n== Normalizacion de URLs de Facebook ==")
    from src.redes.facebook import ExtractorFacebook

    fb = ExtractorFacebook()
    n = fb._normalizar_url_publicacion

    comprobar(
        n("https://www.facebook.com/pagina/posts/123?__cft__[0]=abc&__tn__=x")
        == "https://www.facebook.com/pagina/posts/123",
        "Quita los parametros de seguimiento",
    )
    comprobar(
        n("https://www.facebook.com/permalink.php?story_fbid=99&id=55&__cft__[0]=z")
        == "https://www.facebook.com/permalink.php?story_fbid=99&id=55",
        "Conserva story_fbid e id",
    )
    comprobar(
        n("https://www.facebook.com/photo/?fbid=777&set=a.123")
        == "https://www.facebook.com/photo?fbid=777",
        "Conserva solo fbid en las fotos",
    )
    comprobar(n("https://www.facebook.com/login/?next=x") == "", "Descarta /login")
    comprobar(n("https://ejemplo.com/algo") == "", "Descarta enlaces de otros sitios")

    comprobar(fb._tipo_publicacion("https://www.facebook.com/reel/9") == "reel", "Tipo reel")
    comprobar(fb._tipo_publicacion("https://www.facebook.com/p/videos/9") == "video", "Tipo video")
    comprobar(
        fb._tipo_publicacion("https://www.facebook.com/photo?fbid=1") == "foto", "Tipo foto"
    )

    print("\n== Normalizacion de URL de perfil ==")
    comprobar(
        fb.normalizar_url_perfil("nombredepagina")
        == "https://www.facebook.com/nombredepagina",
        "Solo el nombre -> URL completa",
    )
    comprobar(
        fb.normalizar_url_perfil("@nombredepagina")
        == "https://www.facebook.com/nombredepagina",
        "Con @ delante",
    )
    comprobar(
        fb.normalizar_url_perfil("https://m.facebook.com/pagina/")
        == "https://www.facebook.com/pagina/",
        "m.facebook -> www.facebook",
    )
    comprobar(
        "profile.php?id=100" in fb.normalizar_url_perfil(
            "https://www.facebook.com/profile.php?id=100"
        ),
        "Conserva profile.php?id=",
    )

    print("\n== Identificador de comentario (anti duplicados) ==")
    c1 = Comentario("u", "Ana", "Hola")
    c2 = Comentario("u", "Ana", "Hola")
    c3 = Comentario("u", "Ana", "Hola!")
    comprobar(c1.id == c2.id, "Mismo comentario -> mismo id")
    comprobar(c1.id != c3.id, "Comentario distinto -> id distinto")


def main() -> int:
    print("=" * 62)
    print("  PRUEBAS DE LA LOGICA INTERNA")
    print("=" * 62)
    probar_fechas()
    probar_limpieza()
    carpeta = probar_base_datos()
    probar_facebook_urls()

    print("\n" + "=" * 62)
    if fallos:
        print(f"  {len(fallos)} PRUEBA(S) FALLARON:")
        for f in fallos:
            print(f"   - {f}")
        return 1
    print("  TODAS LAS PRUEBAS PASARON ✔")
    print(f"  (archivos temporales en {carpeta})")
    print("=" * 62)
    return 0


if __name__ == "__main__":
    sys.exit(main())
